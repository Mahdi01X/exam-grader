"""Exports : tableur classe (xlsx) et PDF annoté par copie."""
from __future__ import annotations
import io
from typing import Iterable
from xml.sax.saxutils import escape as _xml_escape

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle,
    PageBreak,
)
from reportlab.lib import colors

from app.models.copy import StudentCopy
from app.models.exam import Exam
from app.models.grade import QuestionGrade
from app.models.rubric import RubricItem


def _esc(text: object) -> str:
    """Échappe le texte dynamique inséré dans un Paragraph ReportLab.

    ReportLab interprète `<`, `>` et `&` comme du balisage : sans échappement,
    une réponse contenant « x < 5 » ou « R&D » casse le rendu (ou affiche une
    balise brute). On échappe donc tout contenu venant de la BD / de l'IA.
    """
    return _xml_escape(str(text if text is not None else ""))


def class_xlsx(exam: Exam, rubric: list[RubricItem], copies: list[StudentCopy],
               grades_by_copy: dict[int, list[QuestionGrade]]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Notes"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1E40AF")

    headers = ["Identifiant"] + [f"Q{r.question_number} / {r.points_max}" for r in rubric] + ["Total", "Total max", "Statut"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill

    total_max = sum(r.points_max for r in rubric)

    for row, copy in enumerate(copies, start=2):
        ws.cell(row=row, column=1, value=copy.student_identifier)
        grades = {g.rubric_item_id: g for g in grades_by_copy.get(copy.id, [])}
        total = 0.0
        for col, r in enumerate(rubric, start=2):
            g = grades.get(r.id)
            pts = (g.final_points if g and g.final_points is not None else (g.proposed_points if g else 0.0))
            ws.cell(row=row, column=col, value=round(pts, 2))
            total += pts
        ws.cell(row=row, column=len(headers) - 2, value=round(total, 2))
        ws.cell(row=row, column=len(headers) - 1, value=round(total_max, 2))
        ws.cell(row=row, column=len(headers), value=copy.status.value)

    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def annotated_pdf(exam: Exam, copy: StudentCopy, rubric: list[RubricItem],
                  grades: list[QuestionGrade]) -> bytes:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=1.5 * cm, rightMargin=1.5 * cm)
    styles = getSampleStyleSheet()
    # Style compact pour les cellules. Un Paragraph s'enroule dans la largeur de
    # la colonne ; une chaîne brute, elle, déborde sur la colonne voisine — c'est
    # ce qui faisait chevaucher « Règle » et « Justification ».
    cell = ParagraphStyle("cell", parent=styles["BodyText"], fontSize=8, leading=10)
    story = []

    story.append(Paragraph(_esc(exam.title), styles["Title"]))
    story.append(Paragraph(f"Matière : {_esc(exam.subject or '—')}", styles["Normal"]))
    story.append(Paragraph(f"Étudiant : <b>{_esc(copy.student_identifier)}</b>", styles["Normal"]))
    story.append(Spacer(1, 0.5 * cm))

    grades_by_q = {g.rubric_item_id: g for g in grades}
    total_pts = 0.0
    total_max = 0.0

    data = [["Question", "Points", "Règle", "Justification", "Conf."]]
    for r in rubric:
        g = grades_by_q.get(r.id)
        pts = g.final_points if g and g.final_points is not None else (g.proposed_points if g else 0.0)
        total_pts += pts
        total_max += r.points_max
        policy_name = g.applied_policy.name if g and g.applied_policy else "—"
        justification = (g.justification if g else "—")[:240]
        confidence = f"{g.confidence:.2f}" if g else "—"
        data.append([
            Paragraph(f"<b>Q{_esc(r.question_number)}</b><br/>{_esc(r.intitule[:120])}", cell),
            f"{pts:.2f} / {r.points_max:.2f}",
            Paragraph(_esc(policy_name), cell),          # Paragraph => le texte s'enroule
            Paragraph(_esc(justification), cell),
            confidence,
        ])
    # Ligne total : texte simple. Le gras vient du TableStyle ci-dessous, PAS de
    # balises HTML (une chaîne de table n'est pas interprétée : « <b> » s'afficherait
    # tel quel, comme dans l'ancien export).
    data.append(["", f"{total_pts:.2f} / {total_max:.2f}", "", "", ""])

    table = Table(
        data,
        colWidths=[3.7 * cm, 2.3 * cm, 4.2 * cm, 5.8 * cm, 1.5 * cm],
        repeatRows=1,
    )
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E40AF")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 8),
        ("GRID", (0, 0), (-1, -2), 0.25, colors.grey),
        ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#F1F5F9")),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, -1), (-1, -1), 9),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        # Colonnes en texte brut (Points, Conf.) : taille 8 pour rester cohérent.
        ("FONTSIZE", (1, 1), (1, -2), 8),
        ("FONTSIZE", (4, 1), (4, -2), 8),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    story.append(table)

    # Transcriptions détaillées par question
    story.append(PageBreak())
    story.append(Paragraph("Détail des transcriptions", styles["Heading2"]))
    for r in rubric:
        g = grades_by_q.get(r.id)
        story.append(
            Paragraph(f"Q{_esc(r.question_number)} — {_esc(r.intitule[:200])}", styles["Heading3"])
        )
        story.append(Paragraph(f"<i>Attendu :</i> {_esc(r.expected_answer[:500])}", styles["BodyText"]))
        if g:
            story.append(
                Paragraph(f"<i>Transcription :</i> {_esc((g.extracted_text or '—')[:1000])}", styles["BodyText"])
            )
            story.append(
                Paragraph(f"<i>Justification :</i> {_esc(g.justification or '—')}", styles["BodyText"])
            )
        story.append(Spacer(1, 0.2 * cm))

    doc.build(story)
    return buf.getvalue()
